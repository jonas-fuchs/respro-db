#!/usr/bin/env python3
"""
Converter for HSV Drug Resistance Database (Dähne et al. 2025, Zenodo 15149867).

Downloads the Excel supplementary table, converts it to ResPro-compatible TSV files,
and writes metadata.json. The downloaded file is removed after extraction.

Usage:
    python convert.py [--output-dir <path>]
"""

import argparse
import hashlib
import json
import re
import sys
import tempfile
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ZENODO_URL = (
    "https://zenodo.org/records/15149867/files/"
    "Supplementary%20table%203.docx.xlsx"
)

GENE_MAP = {
    "Thymidine Kinase": "UL23",
    "DNA Polymerase": "UL30",
    "UL5": "UL5",
    "UL52": "UL52",
}

REFERENCE_IDS = {
    "HSV-1": "NC_001806",
    "HSV-2": "NC_001798",
}

# Sheets grouped by virus type
SHEET_VIRUS = {
    "HSV-1 TK": "HSV-1",
    "HSV-1 Pol": "HSV-1",
    "HSV-1 HPC": "HSV-1",
    "HSV-2 TK": "HSV-2",
    "HSV-2 Pol": "HSV-2",
    "HSV-2 HPC": "HSV-2",
}

RULES_COLUMNS = [
    "gene",
    "reference_identifier",
    "position",
    "reference",
    "mutation",
    "antiviral",
    "group_id",
    "member_id",
    "phenotype",
    "clinical_phenotype",
    "publication",
    "source",
    "comment",
]

FORMULA_COLUMNS = [
    "group_id",
    "antiviral",
    "expression",
    "phenotype",
    "clinical_phenotype",
    "publication",
    "source",
]

NON_MIGRATED_COLUMNS = [
    "reason",
    "sheet",
    "gene",
    "drug",
    "aa_change",
    "aa_position",
    "resistance",
    "pmids",
    "details",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def norm_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def count_pmids(pmid_str: str) -> int:
    if not pmid_str:
        return 0
    return len([p for p in re.split(r"[\s,;]+", pmid_str) if p.strip()])


def parse_pmids(pmid_str: str) -> str:
    """Normalise PMIDs to a comma-separated list with PMID: prefix."""
    if not pmid_str:
        return ""
    # Source rows may separate PMIDs with commas, semicolons, or whitespace.
    raw_parts = [p.strip() for p in re.split(r"[\s,;]+", pmid_str) if p.strip()]
    norm_parts = []
    for part in raw_parts:
        cleaned = re.sub(r"^PMID:\s*", "", part, flags=re.IGNORECASE)
        norm_parts.append(f"PMID:{cleaned}")
    return ",".join(norm_parts)


def source_website_from_url(source_url: str) -> str:
    match = re.search(r"/records/(\d+)/", source_url)
    if match:
        return f"https://zenodo.org/records/{match.group(1)}"
    return "https://zenodo.org/records/15149867"


def zenodo_record_id_from_url(source_url: str) -> str:
    match = re.search(r"/records/(\d+)/", source_url)
    if not match:
        raise ValueError(f"Could not determine Zenodo record id from source URL: {source_url}")
    return match.group(1)


def zenodo_record_metadata(record_id: str) -> dict:
    api_url = f"https://zenodo.org/api/records/{record_id}"
    with urllib.request.urlopen(api_url) as response:
        return json.load(response)


def zenodo_record_update_date(record_metadata: dict) -> str:
    for key in ("updated", "created"):
        value = str(record_metadata.get(key, "")).strip()
        if value:
            return value.split("T", 1)[0]
    raise ValueError("Zenodo record metadata did not contain created/updated timestamp")


def parse_aa_change(aa_change: str):
    """
    Return (reference_aa, position, mutation_token) from an AA change string.

    Supported forms:
      A336V          — substitution
      A336*          — stop
      N23Stop        — stop with ref AA
      8Stop          — stop without ref AA
      Y101fsx        — frameshift with ref AA
      144fsx         — frameshift without ref AA
      I194Del        — single-residue deletion
      1-248Del       — range deletion (no leading AAs)
      DD676-677Del   — multi-residue range deletion
      V813M*         — substitution with trailing asterisk annotation (strip *)

    Returns (ref_aa_or_None, position_int_or_None, mutation_token_str).
    Returns (None, None, None) when the mutation cannot be represented.
    """
    aa_change = aa_change.strip().replace("\n", "")

    # Dual allele like "E43A/D" — ambiguous, skip
    if re.match(r"^[A-Z]\d+[A-Z]/[A-Z]$", aa_change, re.IGNORECASE):
        return None, None, None

    # Strip spurious trailing asterisk that is not itself a stop notation
    # e.g. V813M* where * is an annotation marker
    trailing_star = re.match(r"^([A-Z])(\d+)([A-Z])\*$", aa_change, re.IGNORECASE)
    if trailing_star:
        ref_aa = trailing_star.group(1).upper()
        pos = int(trailing_star.group(2))
        alt_aa = trailing_star.group(3).upper()
        return ref_aa, pos, alt_aa

    # frameshift with ref AA: e.g. Y101fsx, K715fsATFF*, Y101fs
    fs_match = re.match(r"^([A-Z])(\d+)(fs.*)$", aa_change, re.IGNORECASE)
    if fs_match:
        ref_aa = fs_match.group(1).upper()
        pos = int(fs_match.group(2))
        return ref_aa, pos, f"{ref_aa}fsX"

    # frameshift without ref AA: e.g. 144fsx, 145fs
    fs_nore = re.match(r"^(\d+)(fs.*)$", aa_change, re.IGNORECASE)
    if fs_nore:
        pos = int(fs_nore.group(1))
        return None, pos, "fsX"

    # standard substitution / stop: A336V, A336*
    sub_match = re.match(r"^([A-Z])(\d+)([A-Z\*])$", aa_change, re.IGNORECASE)
    if sub_match:
        ref_aa = sub_match.group(1).upper()
        pos = int(sub_match.group(2))
        alt_aa = sub_match.group(3).upper()
        if alt_aa == "*":
            return ref_aa, pos, "*"
        return ref_aa, pos, alt_aa

    # stop with ref AA: N23Stop, M182Stop
    stop_re = re.match(r"^([A-Z])(\d+)[Ss]top$", aa_change, re.IGNORECASE)
    if stop_re:
        ref_aa = stop_re.group(1).upper()
        pos = int(stop_re.group(2))
        return ref_aa, pos, "*"

    # stop without ref AA: 8Stop, 44Stop
    stop_nore = re.match(r"^(\d+)[Ss]top$", aa_change, re.IGNORECASE)
    if stop_nore:
        pos = int(stop_nore.group(1))
        return None, pos, "*"

    # single-residue deletion: e.g. "I194Del"
    single_del = re.match(r"^([A-Z])(\d+)[Dd]el$", aa_change, re.IGNORECASE)
    if single_del:
        ref_aa = single_del.group(1).upper()
        pos = int(single_del.group(2))
        return ref_aa, pos, f"{ref_aa}{pos}del"

    # Single-position insertion without range: e.g. "A301Ins", "684Ins"
    # Cannot determine inserted sequence — skip
    if re.match(r"^[A-Z]?\d+[Ii]ns$", aa_change, re.IGNORECASE):
        return None, None, None

    # Multi-residue range deletion/insertion with leading AAs:
    # e.g. DD676-677Del, PGDEPA1106-1111Del, ED684-685Ins
    multi_indel = re.match(
        r"^([A-Z]+)(\d+)-(\d+)(Del|Ins)$", aa_change, re.IGNORECASE
    )
    if multi_indel:
        ref_aas = multi_indel.group(1).upper()
        start = int(multi_indel.group(2))
        kind = multi_indel.group(4).lower()
        ref_aa = ref_aas[0]
        if kind == "del":
            return ref_aa, start, f"{ref_aas}{start}del"
        else:
            return ref_aa, start, f"{ref_aa}{start}ins{ref_aas[1:]}"

    # bare range deletion without leading AAs: e.g. 1-248Del
    bare_del = re.match(r"^(\d+)-(\d+)[Dd]el$", aa_change)
    if bare_del:
        start = int(bare_del.group(1))
        return None, start, f"del{start}_{bare_del.group(2)}"

    # unrecognised
    return None, None, None


def determine_phenotypes(resistance: str, cell_culture: str, clinical: str):
    """
    Returns (phenotype, clinical_phenotype) according to mapping rules.
    """

    def define_phenotype(resistance, cc_cl):
        if resistance == "sensitive":
            if cc_cl == "yes":
                return "sensitive"
            elif cc_cl == "contradiction":
                return "resistant"
            else:
                return "unknown"
        if resistance == "resistant":
            if cc_cl == "yes":
                return "resistant"
            elif cc_cl == "contradiction":
                return "sensitive"
            else:
                return "unknown"
        if resistance == "sensitive/resistant":
            if cc_cl == "yes":
                return "contradictory"
            elif cc_cl == "contradiction":
                return "contradictory"
            else:
                return "unknown"
        
        return "unknown"

    r = resistance.strip().lower()
    cc = cell_culture.strip().lower()
    cl = clinical.strip().lower()

    return define_phenotype(r, cc), define_phenotype(r, cl)


def tsv_checksum(content: str) -> str:
    return "sha256:" + hashlib.sha256(content.encode("utf-8")).hexdigest()


def add_non_migrated(
    non_migrated_rows,
    reason: str,
    sheet: str,
    gene: str,
    drug: str,
    aa_change: str,
    aa_position,
    resistance: str,
    pmids: str,
    details: str = "",
) -> None:
    non_migrated_rows.append(
        {
            "reason": reason,
            "sheet": sheet,
            "gene": gene,
            "drug": drug,
            "aa_change": aa_change,
            "aa_position": "" if aa_position is None else str(aa_position),
            "resistance": resistance,
            "pmids": pmids,
            "details": details,
        }
    )


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------


def extract_rows(wb):
    """
    Extract all relevant rows from the workbook.

    Returns:
        single_rows: list of dicts for rules.tsv (non-combo rows + combo member rows)
        formula_rows: list of dicts for formula-rules.tsv
    """
    single_rows = []
    formula_rows = []
    non_migrated_rows = []
    group_counter = [0]

    def next_group_id():
        group_counter[0] += 1
        return f"G{group_counter[0]:04d}"

    for sheet_name, virus in SHEET_VIRUS.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: Sheet '{sheet_name}' not found, skipping.", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        ref_id = REFERENCE_IDS[virus]

        for row in ws.iter_rows(min_row=2, values_only=True):
            gene_raw = norm_str(row[0])
            if not gene_raw:
                continue

            drug = norm_str(row[1]).strip()
            aa_change = norm_str(row[2])
            aa_pos_raw = row[3]
            resistance = norm_str(row[4])
            pmids_raw = norm_str(row[5])
            cell_culture = norm_str(row[6])
            clinical = norm_str(row[7])
            comment = norm_str(row[10]) if len(row) > 10 else ""

            pmids = parse_pmids(pmids_raw)

            # Skip rows where AA change is absent (nucleotide-only entries)
            if not aa_change:
                print(
                    f"SKIPPED nucleotide-only row: sheet={sheet_name} "
                    f"pos_col={aa_pos_raw!r} drug={drug} "
                    f"(no amino-acid change recorded)",
                    file=sys.stderr,
                )
                add_non_migrated(
                    non_migrated_rows,
                    reason="nucleotide_only_row",
                    sheet=sheet_name,
                    gene=gene_raw,
                    drug=drug,
                    aa_change=aa_change,
                    aa_position=aa_pos_raw,
                    resistance=resistance,
                    pmids=pmids,
                    details="No amino-acid change value provided in source row.",
                )
                continue

            # Slash-separated mutations define a combo rule.
            # If genes are slash-separated too, members are matched by index.
            # If only one gene is provided, it applies to all mutation members.
            parts_aa = [part.strip() for part in aa_change.split("/") if part.strip()]
            is_dual_allele = bool(re.match(r"^[A-Z]\d+[A-Z]/[A-Z]$", aa_change, re.IGNORECASE))
            is_combo = len(parts_aa) > 1 and not is_dual_allele
            if is_combo:
                raw_parts_gene = [part.strip() for part in gene_raw.split("/") if part.strip()]
                if len(raw_parts_gene) == 1:
                    parts_gene = raw_parts_gene * len(parts_aa)
                else:
                    parts_gene = raw_parts_gene

                # Position column may also be slash-split ("356/222").
                if aa_pos_raw and "/" in str(aa_pos_raw):
                    parts_pos = [part.strip() for part in str(aa_pos_raw).split("/") if part.strip()]
                else:
                    parts_pos = [str(aa_pos_raw)] * len(parts_aa)

                if len(parts_gene) != len(parts_aa) or len(parts_pos) != len(parts_aa):
                    print(
                        f"WARNING: Cannot parse combo row gene='{gene_raw}' "
                        f"aa='{aa_change}' in sheet '{sheet_name}' — skipping.",
                        file=sys.stderr,
                    )
                    add_non_migrated(
                        non_migrated_rows,
                        reason="invalid_combo_format",
                        sheet=sheet_name,
                        gene=gene_raw,
                        drug=drug,
                        aa_change=aa_change,
                        aa_position=aa_pos_raw,
                        resistance=resistance,
                        pmids=pmids,
                        details=(
                            "Expected the same number of mutation and position members, and "
                            "either one gene or one gene per mutation member."
                        ),
                    )
                    continue

                gid = next_group_id()
                member_ids = []

                for sub_gene, sub_aa, sub_pos_raw in zip(parts_gene, parts_aa, parts_pos):
                    mapped_gene = GENE_MAP.get(sub_gene, sub_gene)
                    ref_aa, pos, mut_token = parse_aa_change(sub_aa)
                    if mut_token is None or ref_aa is None:
                        print(
                            f"WARNING: Cannot parse mutation '{sub_aa}' (or missing ref AA) in combo row "
                            f"(sheet={sheet_name}, gene={sub_gene}) — skipping entire combo.",
                            file=sys.stderr,
                        )
                        add_non_migrated(
                            non_migrated_rows,
                            reason="combo_member_unparseable_or_missing_ref",
                            sheet=sheet_name,
                            gene=sub_gene,
                            drug=drug,
                            aa_change=sub_aa,
                            aa_position=sub_pos_raw,
                            resistance=resistance,
                            pmids=pmids,
                            details="Combo row dropped because one member could not be represented.",
                        )
                        member_ids = []
                        break
                    if pos is None:
                        try:
                            pos = int(sub_pos_raw)
                        except (ValueError, TypeError):
                            pos = 0

                    mid = f"{gid}_{mapped_gene}_{sub_aa}"
                    member_ids.append(mid)

                    # combo members: no phenotype per user instructions
                    single_rows.append({
                        "gene": mapped_gene,
                        "reference_identifier": ref_id,
                        "position": pos,
                        "reference": ref_aa or "",
                        "mutation": mut_token,
                        "antiviral": drug.lower(),
                        "group_id": gid,
                        "member_id": mid,
                        "phenotype": "",
                        "clinical_phenotype": "",
                        "publication": pmids,
                        "source": "Dähne et al. 2025 (Zenodo 15149867)",
                        "comment": comment,
                    })

                if len(member_ids) != len(parts_gene):
                    # Already logged above; remove any partially-added members
                    single_rows = [r for r in single_rows if r.get("group_id") != gid]
                    continue

                # formula row
                phenotype, clinical_phenotype = determine_phenotypes(
                    resistance, cell_culture, clinical
                )
                expr = "(" + " AND ".join(member_ids) + ")"
                formula_rows.append({
                    "group_id": gid,
                    "antiviral": drug.lower(),
                    "expression": expr,
                    "phenotype": phenotype,
                    "clinical_phenotype": clinical_phenotype,
                    "publication": pmids,
                    "source": "Dähne et al. 2025 (Zenodo 15149867)",
                })

            else:
                # Regular single-gene row
                mapped_gene = GENE_MAP.get(gene_raw)
                if mapped_gene is None:
                    print(
                        f"WARNING: Unknown gene '{gene_raw}' in sheet '{sheet_name}' — skipping.",
                        file=sys.stderr,
                    )
                    add_non_migrated(
                        non_migrated_rows,
                        reason="unknown_gene",
                        sheet=sheet_name,
                        gene=gene_raw,
                        drug=drug,
                        aa_change=aa_change,
                        aa_position=aa_pos_raw,
                        resistance=resistance,
                        pmids=pmids,
                        details="Gene not mappable to ResPro gene naming.",
                    )
                    continue

                ref_aa, pos, mut_token = parse_aa_change(aa_change)
                if mut_token is None:
                    print(
                        f"SKIPPED unparseable mutation: sheet={sheet_name} "
                        f"gene={mapped_gene} aa_change={aa_change!r} drug={drug}",
                        file=sys.stderr,
                    )
                    add_non_migrated(
                        non_migrated_rows,
                        reason="unparseable_mutation",
                        sheet=sheet_name,
                        gene=mapped_gene,
                        drug=drug,
                        aa_change=aa_change,
                        aa_position=aa_pos_raw,
                        resistance=resistance,
                        pmids=pmids,
                        details="Mutation syntax could not be converted to supported ResPro notation.",
                    )
                    continue
                if ref_aa is None:
                    print(
                        f"SKIPPED mutation with no reference AA: sheet={sheet_name} "
                        f"gene={mapped_gene} aa_change={aa_change!r} drug={drug} "
                        f"(reference amino acid required by ResPro)",
                        file=sys.stderr,
                    )
                    add_non_migrated(
                        non_migrated_rows,
                        reason="missing_reference_amino_acid",
                        sheet=sheet_name,
                        gene=mapped_gene,
                        drug=drug,
                        aa_change=aa_change,
                        aa_position=aa_pos_raw,
                        resistance=resistance,
                        pmids=pmids,
                        details="Reference amino acid is required by rules.tsv schema.",
                    )
                    continue
                if pos is None:
                    try:
                        pos = int(aa_pos_raw)
                    except (ValueError, TypeError):
                        print(
                            f"SKIPPED row with unresolvable position: sheet={sheet_name} "
                            f"gene={mapped_gene} aa_change={aa_change!r} pos_col={aa_pos_raw!r}",
                            file=sys.stderr,
                        )
                        add_non_migrated(
                            non_migrated_rows,
                            reason="unresolvable_position",
                            sheet=sheet_name,
                            gene=mapped_gene,
                            drug=drug,
                            aa_change=aa_change,
                            aa_position=aa_pos_raw,
                            resistance=resistance,
                            pmids=pmids,
                            details="Could not resolve amino-acid position to integer.",
                        )
                        continue

                phenotype, clinical_phenotype = determine_phenotypes(
                    resistance, cell_culture, clinical
                )

                single_rows.append({
                    "gene": mapped_gene,
                    "reference_identifier": ref_id,
                    "position": pos,
                    "reference": ref_aa or "",
                    "mutation": mut_token,
                    "antiviral": drug.lower(),
                    "group_id": "",
                    "member_id": "",
                    "phenotype": phenotype,
                    "clinical_phenotype": clinical_phenotype,
                    "publication": pmids,
                    "source": "Dähne et al. 2025 (Zenodo 15149867)",
                    "comment": comment,
                })

    return single_rows, formula_rows, non_migrated_rows


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------


def deduplicate_single_rows(rows, non_migrated_rows):
    """
    For rows with identical (gene, reference_identifier, position, mutation, antiviral)
    keep the one with more PMIDs. Print dropped rows.
    """
    # Only deduplicate non-combo rows (member_id == "")
    combo = [r for r in rows if r["member_id"]]
    non_combo = [r for r in rows if not r["member_id"]]

    key_map = defaultdict(list)
    for r in non_combo:
        key = (r["gene"], r["reference_identifier"], r["position"], r["mutation"], r["antiviral"])
        key_map[key].append(r)

    kept = []
    for key, group in key_map.items():
        if len(group) == 1:
            kept.append(group[0])
        else:
            best = max(group, key=lambda r: count_pmids(r["publication"]))
            for r in group:
                if r is not best:
                    print(
                        f"DROPPED duplicate row: gene={r['gene']} pos={r['position']} "
                        f"mutation={r['mutation']} drug={r['antiviral']} "
                        f"pmids={r['publication']!r} "
                        f"(kept pmids={best['publication']!r}, "
                        f"reason: fewer PMIDs than retained row)",
                        file=sys.stderr,
                    )
                    add_non_migrated(
                        non_migrated_rows,
                        reason="duplicate_rule_dropped_fewer_pmids",
                        sheet="",
                        gene=r["gene"],
                        drug=r["antiviral"],
                        aa_change=f"{r['reference']}{r['position']}{r['mutation']}",
                        aa_position=r["position"],
                        resistance="",
                        pmids=r["publication"],
                        details=(
                            f"Kept publication set: {best['publication']}; "
                            "this row had fewer PMID entries."
                        ),
                    )
            kept.append(best)

    return kept + combo


# ---------------------------------------------------------------------------
# Serialisation
# ---------------------------------------------------------------------------


def rows_to_tsv(rows, columns) -> str:
    lines = ["\t".join(columns)]
    for r in rows:
        lines.append("\t".join(str(r.get(c, "")) for c in columns))
    return "\n".join(lines) + "\n"


def sort_single_rows(rows):
    return sorted(
        rows,
        key=lambda r: (
            r["reference_identifier"],
            r["gene"],
            r["antiviral"],
            int(r["position"]) if str(r["position"]).isdigit() else 0,
            r["mutation"],
        ),
    )


def sort_formula_rows(rows):
    return sorted(rows, key=lambda r: (r["group_id"], r["antiviral"]))


def non_migrated_rows_to_text(rows) -> str:
    lines = [
        "# Non-migrated rules from source workbook",
        "# Columns: " + "\t".join(NON_MIGRATED_COLUMNS),
        "",
        "\t".join(NON_MIGRATED_COLUMNS),
    ]
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            r["reason"],
            r["sheet"],
            r["gene"],
            r["drug"],
            r["aa_change"],
        ),
    )
    for row in sorted_rows:
        lines.append("\t".join(str(row.get(col, "")) for col in NON_MIGRATED_COLUMNS))
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Convert HSV resistance DB from Zenodo.")
    parser.add_argument(
        "--source-url",
        default=ZENODO_URL,
        help="Direct source file URL (defaults to Zenodo record v3 XLSX).",
    )
    parser.add_argument(
        "--output-dir",
        default=str(Path(__file__).resolve().parent.parent / "output"),
        help="Directory to write output files into.",
    )
    args = parser.parse_args()

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    record_id = zenodo_record_id_from_url(args.source_url)
    record_metadata = zenodo_record_metadata(record_id)
    maintainer_update = zenodo_record_update_date(record_metadata)

    # --- Download ---
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_path = Path(tmp_file.name)
    print(f"Downloading {args.source_url} …", file=sys.stderr)
    urllib.request.urlretrieve(args.source_url, tmp_path)
    print(f"Download complete ({tmp_path.stat().st_size} bytes).", file=sys.stderr)

    try:
        wb = openpyxl.load_workbook(tmp_path)
        single_rows, formula_rows, non_migrated_rows = extract_rows(wb)
    finally:
        tmp_path.unlink(missing_ok=True)
        print("Temporary file deleted.", file=sys.stderr)

    # --- Deduplication ---
    single_rows = deduplicate_single_rows(single_rows, non_migrated_rows)

    # --- Sort ---
    single_rows = sort_single_rows(single_rows)
    formula_rows = sort_formula_rows(formula_rows)

    # --- Write rules.tsv ---
    rules_content = rows_to_tsv(single_rows, RULES_COLUMNS)
    rules_path = out_dir / "rules.tsv"
    rules_path.write_text(rules_content, encoding="utf-8")
    print(f"Written {rules_path} ({len(single_rows)} rows).", file=sys.stderr)

    # --- Write formula-rules.tsv (only if combo rows exist) ---
    if formula_rows:
        formula_content = rows_to_tsv(formula_rows, FORMULA_COLUMNS)
        formula_path = out_dir / "formula-rules.tsv"
        formula_path.write_text(formula_content, encoding="utf-8")
        print(f"Written {formula_path} ({len(formula_rows)} rows).", file=sys.stderr)
    else:
        formula_content = ""

    # --- Write metadata.json ---
    checksum = tsv_checksum(rules_content)
    metadata = {
        "maintainers": ["Dähne, Theo", "Gosert, Rainer", "Jaki, Lena"],
        "contact": "marcus.panning@uniklinik-freiburg.de",
        "publication_pmid": "40349973",
        "website": source_website_from_url(args.source_url),
        "description": (
            "HSV-1 and HSV-2 drug resistance mutation database curated from published literature. "
            "Covers TK (UL23), DNA Polymerase (UL30), and Helicase-Primase Complex (UL5/UL52) genes. "
        ),
        "maintainer_update": maintainer_update,
        "license": "CC-BY-4.0",
        "tsv_checksum": checksum,
    }
    metadata_path = out_dir / "metadata.json"
    metadata_path.write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")
    print(f"Written {metadata_path}.", file=sys.stderr)

    # --- Write non-migrated-rules.txt ---
    non_migrated_path = out_dir / "non-migrated-rules.txt"
    non_migrated_content = non_migrated_rows_to_text(non_migrated_rows)
    non_migrated_path.write_text(non_migrated_content, encoding="utf-8")
    print(
        f"Written {non_migrated_path} ({len(non_migrated_rows)} aggregated entries).",
        file=sys.stderr,
    )

    print("Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
